#!/usr/bin/env python3
"""
SX_Steady_State_Model.xlsx  —  Professional single-sheet version
Color system:
  INPUT  = pale yellow  #FFFDE7 / navy font    — hard-coded values the user edits
  CALC   = white        #FFFFFF  / black font   — intermediate formula cells
  STEP   = pale blue    #E8EAF6 / black font    — TDMA forward-sweep intermediate
  OUTPUT = pale green   #F1F8E9 / dark-green    — solution & final results
  REF    = pale lavender #EDE7F6 / black        — cells that pull from elsewhere
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter as gcl

# ── Layout constants ──────────────────────────────────────────────────────────
N       = 25
R0      = 20           # first SOLVER data row (stage 1)
R1      = R0 + N - 1  # last  SOLVER data row (stage 25 = row 44)
PROF_HDR = R1 + 3      # stage-profile header row = 47
PROF_R0  = PROF_HDR + 1
PROF_R1  = PROF_R0 + N - 1

ELEMS  = ["Pr", "Nd", "Tb", "Dy"]
ESTART = {"Pr": 3, "Nd": 11, "Tb": 19, "Dy": 27}
XCOL   = {"Pr": 35, "Nd": 36, "Tb": 37, "Dy": 38}

# ── Cell refs (same sheet, no prefix) ────────────────────────────────────────
REF = {
    "n_ext": "$B$5",  "n_scr": "$B$6",  "n_str": "$B$7",
    "He":    "$B$8",  "Hs":    "$B$9",  "Ht":    "$B$10",
    "F":     "$B$11", "S":     "$B$12",
    "O":     "$B$13", "R":     "$B$14",
    "P_Pr":  "$H$5",  "Q_Pr":  "$I$5",  "XF_Pr": "$M$5",
    "P_Nd":  "$H$6",  "Q_Nd":  "$I$6",  "XF_Nd": "$M$6",
    "P_Tb":  "$H$7",  "Q_Tb":  "$I$7",  "XF_Tb": "$M$7",
    "P_Dy":  "$H$8",  "Q_Dy":  "$I$8",  "XF_Dy": "$M$8",
}

# ── Professional colour system ────────────────────────────────────────────────
# Structure colours
T_TITLE  = "1B3A6B"   # deep navy     — main title
T_SECHDR = "2E4057"   # dark slate    — section headers
T_COLHDR = "4A6FA5"   # steel blue    — column headers
T_LEGEND = "37474F"   # blue-grey     — legend bar

# Semantic cell colours
I_BG = "FFFDE7"  # pale yellow  — INPUT  (hardcoded values)
I_FG = "1A237E"  # dark navy    — INPUT  font
C_BG = "FFFFFF"  # white        — CALC   (formula, intermediate)
C_FG = "212121"  # near-black   — CALC   font
S_BG = "E8EAF6"  # pale indigo  — STEP   (TDMA c′ d′ — intermediate sweep)
S_FG = "212121"  # near-black   — STEP   font
O_BG = "F1F8E9"  # pale mint    — OUTPUT (y_org solution, final results)
O_FG = "2E7D32"  # dark green   — OUTPUT font
R_BG = "E3F2FD"  # pale sky     — REF    (x_aq derived, profile pulls)
R_FG = "0D47A1"  # dark blue    — REF    font
WHT  = "FFFFFF"

# Solver element header colours (muted, professional)
E_CLR = {"Pr": "546E7A",   # blue-grey
          "Nd": "455A64",   # darker blue-grey
          "Tb": "37474F",   # darkest blue-grey
          "Dy": "4E6073"}   # slate

# Stage section tints (very subtle — section identity only)
SEC_EXT = "FFFDE7"   # same as INPUT bg — extraction = loading
SEC_SCR = "F3E5F5"   # pale lavender    — scrub
SEC_STR = "E8F5E9"   # pale mint        — strip = output

# ── Style helpers ─────────────────────────────────────────────────────────────
def fill(c): return PatternFill("solid", fgColor=c)

def font(bold=False, col="212121", sz=10):
    return Font(bold=bold, color=col, size=sz, name="Calibri")

def aln(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def side(w="thin",  col="D0D0D0"): return Side(style=w, color=col)
def side_m(col="9E9E9E"):           return Side(style="medium", color=col)

def brd(col="D0D0D0"):
    s = side(col=col)
    return Border(left=s, right=s, top=s, bottom=s)

def brd_section():
    m = side_m(); t = side()
    return Border(left=m, right=m, top=t, bottom=t)

def style_input(c, val=None, frm=None, nf=None, h="center"):
    c.value = frm if frm is not None else val
    c.font  = font(col=I_FG)
    c.fill  = fill(I_BG)
    c.alignment = aln(h=h)
    c.border = brd()
    if nf: c.number_format = nf

def style_calc(c, frm, nf=None, h="center", sz=9):
    c.value = frm
    c.font  = font(col=C_FG, sz=sz)
    c.fill  = fill(C_BG)
    c.alignment = aln(h=h)
    c.border = brd()
    if nf: c.number_format = nf

def style_step(c, frm, nf=None):
    c.value = frm
    c.font  = font(col=S_FG, sz=9)
    c.fill  = fill(S_BG)
    c.alignment = aln()
    c.border = brd()
    if nf: c.number_format = nf

def style_output(c, frm=None, val=None, nf=None, h="center", sz=10):
    c.value = frm if frm is not None else val
    c.font  = font(bold=True, col=O_FG, sz=sz)
    c.fill  = fill(O_BG)
    c.alignment = aln(h=h)
    c.border = brd()
    if nf: c.number_format = nf

def style_ref(c, frm, nf=None, h="center"):
    c.value = frm
    c.font  = font(col=R_FG, sz=9)
    c.fill  = fill(R_BG)
    c.alignment = aln(h=h)
    c.border = brd()
    if nf: c.number_format = nf

def hdr_main(ws, row, col1, col2, text, bg=T_TITLE, sz=11):
    ws.merge_cells(f"{gcl(col1)}{row}:{gcl(col2)}{row}")
    c = ws.cell(row=row, column=col1)
    c.value = text
    c.font  = font(bold=True, col=WHT, sz=sz)
    c.fill  = fill(bg)
    c.alignment = aln(h="left")
    return c

def hdr_col(ws, row, col, text, bg=T_COLHDR, sz=9, wrap=False):
    c = ws.cell(row=row, column=col)
    c.value = text
    c.font  = font(bold=True, col=WHT, sz=sz)
    c.fill  = fill(bg)
    c.alignment = aln(wrap=wrap)
    c.border = brd("AAAAAA")
    return c

def label(ws, row, col, text, bg=None, fc="212121", bold=False, h="left", sz=10):
    c = ws.cell(row=row, column=col)
    c.value = text
    c.font  = font(bold=bold, col=fc, sz=sz)
    if bg: c.fill = fill(bg)
    c.alignment = aln(h=h)
    c.border = brd()
    return c

# ── TDMA formula builders ─────────────────────────────────────────────────────
def f_D(row, elem):
    P,Q   = REF[f"P_{elem}"], REF[f"Q_{elem}"]
    ne,ns = REF["n_ext"], REF["n_scr"]
    He,Hs,Ht = REF["He"], REF["Hs"], REF["Ht"]
    return (f"=IF(A{row}<={ne},{P}*{He}^{Q},"
            f"IF(A{row}<={ne}+{ns},{P}*{Hs}^{Q},{P}*{Ht}^{Q}))")

def f_a(row):
    return f"=IF(A{row}=1,0,{REF['O']})"

def f_b(row, Dc):
    O,S,R,F = REF["O"],REF["S"],REF["R"],REF["F"]
    ne,ns,D = REF["n_ext"],REF["n_scr"],f"{Dc}{row}"
    return (f"=IF(A{row}>{ne}+{ns},-{O}-{S}/{D},"
            f"IF(A{row}>{ne},-{O}-{S}*{R}/{D},-{O}-({S}*{R}+{F})/{D}))")

def f_c(row, Dc):
    S,R,F = REF["S"],REF["R"],REF["F"]
    ne,ns = REF["n_ext"],REF["n_scr"]
    if row == R1: return "=0"
    Dn = f"{Dc}{row+1}"
    return (f"=IF(A{row}>{ne}+{ns},{S}/{Dn},"
            f"IF(A{row}>={ne},{S}*{R}/{Dn},({S}*{R}+{F})/{Dn}))")

def f_d(row, elem):
    XF,F,ne = REF[f"XF_{elem}"],REF["F"],REF["n_ext"]
    return f"=IF(A{row}={ne},-{XF}*{F},0)"

def f_cp(row, bc, cc, ac, cpc):
    if row == R0: return f"={cc}{row}/{bc}{row}"
    w = f"({bc}{row}-{ac}{row}*{cpc}{row-1})"
    return f"={cc}{row}/{w}"

def f_dp(row, bc, dc, ac, cpc, dpc):
    if row == R0: return f"={dc}{row}/{bc}{row}"
    w = f"({bc}{row}-{ac}{row}*{cpc}{row-1})"
    return f"=({dc}{row}-{ac}{row}*{dpc}{row-1})/{w}"

def f_y(row, dpc, cpc, yc):
    if row == R1: return f"={dpc}{row}"
    return f"={dpc}{row}-{cpc}{row}*{yc}{row+1}"

# ═════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "SX Model"

# ── Column widths ─────────────────────────────────────────────────────────────
col_widths = {
    "A":30, "B":12, "C":2,  "D":2,  "E":2,
    "F":10, "G":2,  "H":12, "I":12,
    "J":2,  "K":10, "L":2,  "M":14,
    "N":2,  "O":36, "P":14
}
for c,w in col_widths.items():
    ws.column_dimensions[c].width = w
for e in ELEMS:
    for off in range(8):
        ws.column_dimensions[gcl(ESTART[e]+off)].width = 11
for xc in XCOL.values():
    ws.column_dimensions[gcl(xc)].width = 11
ws.column_dimensions[gcl(39)].width = 12
ws.column_dimensions[gcl(40)].width = 12

# ═════════════════════════════════════════════════════════════════════════════
# ROW 1 — MAIN TITLE
# ═════════════════════════════════════════════════════════════════════════════
ws.merge_cells(f"A1:{gcl(40)}1")
c = ws["A1"]
c.value = "SX STEADY-STATE MODEL  ·  Alind Formulation  ·  D2EHPA / HCl  ·  TDMA Direct Solver"
c.font  = font(bold=True, col=WHT, sz=13)
c.fill  = fill(T_TITLE)
c.alignment = aln()
ws.row_dimensions[1].height = 30

# ROW 2 — colour legend bar
ws.row_dimensions[2].height = 18
legend = [
    (1,  4,  "INPUT — hard-coded value",    I_BG, I_FG),
    (5,  8,  "CALC — intermediate formula", C_BG, C_FG),
    (9,  12, "STEP — TDMA sweep (c′ d′)",   S_BG, S_FG),
    (13, 16, "OUTPUT — solution / result",  O_BG, O_FG),
    (17, 20, "REF — derived / linked",      R_BG, R_FG),
]
for c1, c2, text, bg, fg in legend:
    ws.merge_cells(f"{gcl(c1)}2:{gcl(c2)}2")
    c = ws.cell(row=2, column=c1)
    c.value = text
    c.font  = font(bold=True, col=fg, sz=9)
    c.fill  = fill(bg)
    c.alignment = aln()
    c.border = brd("BBBBBB")

# ═════════════════════════════════════════════════════════════════════════════
# ROWS 4-15 — INPUTS  (three panels + results summary)
# ═════════════════════════════════════════════════════════════════════════════
ws.row_dimensions[3].height = 6

# ── Panel A: General Inputs — cols A-B, rows 4-15 ────────────────────────────
hdr_main(ws, 4, 1, 2, "  GENERAL INPUTS", bg=T_SECHDR, sz=10)
hdr_col(ws, 5, 1, "Parameter", bg=T_COLHDR)
hdr_col(ws, 5, 2, "Value",     bg=T_COLHDR)

gen = [
    ("Extraction Stages  (n_ext)",               10),
    ("Scrub Stages  (n_scr)",                    10),
    ("Strip Stages  (n_str)",                     5),
    ("HCl Normality — Extraction  [N]",         0.50),
    ("HCl Normality — Scrub  [N]",              1.37),
    ("HCl Normality — Strip  [N]",               5.0),
    ("Feed Flowrate  (F)",                         25),
    ("Strip Flowrate  (S)",                        12),
    ("Organic Flow  (O)",                        32.5),
    ("Reflux  (R)  [decimal: 0.70 = 70%]",      0.70),
]
for i, (lbl, val) in enumerate(gen, start=6):
    label(ws, i, 1, lbl, bg=I_BG, fc=I_FG)
    c = ws.cell(row=i, column=2)
    style_input(c, val=val, nf="0.00")
ws.row_dimensions[5].height = 16
# B6=n_ext, B7=n_scr, B8=n_str, B9=He, B10=Hs, B11=Ht, B12=F, B13=S, B14=O, B15=R

# ── Panel B: Distribution Coefficients — cols F-I, rows 4-9 ─────────────────
hdr_main(ws, 4, 6, 9, "  DISTRIBUTION COEFFICIENTS   D = P × [HCl]^Q", bg=T_SECHDR, sz=10)
for col, lbl in [(6,"Element"),(8,"P"),(9,"Q")]:
    hdr_col(ws, 5, col, lbl, bg=T_COLHDR)
dist = [("Pr",0.0008,-1.973),("Nd",0.0031,-2.541),("Tb",0.378,-2.624),("Dy",0.5959,-2.431)]
for i,(elem,P,Q) in enumerate(dist, start=6):
    label(ws, i, 6, elem, bg=I_BG, fc=I_FG, h="center")
    c=ws.cell(row=i,column=8); style_input(c, val=P, nf="0.0000")
    c=ws.cell(row=i,column=9); style_input(c, val=Q, nf="0.000")
# H6=P_Pr,I6=Q_Pr … H9=P_Dy,I9=Q_Dy

# ── Panel C: Feed Concentrations — cols K-M, rows 4-9 ────────────────────────
hdr_main(ws, 4, 11, 13, "  FEED CONCENTRATIONS", bg=T_SECHDR, sz=10)
for col,lbl in [(11,"Element"),(13,"Feed (g/L)")]:
    hdr_col(ws, 5, col, lbl, bg=T_COLHDR)
feed = [("Pr",0.0),("Nd",20.8),("Tb",0.0),("Dy",2.68)]
for i,(elem,val) in enumerate(feed, start=6):
    label(ws, i, 11, elem, bg=I_BG, fc=I_FG, h="center")
    c=ws.cell(row=i,column=13); style_input(c, val=val, nf="0.00")
# M6=XF_Pr, M7=XF_Nd, M8=XF_Tb, M9=XF_Dy

# ── Panel D: Results Summary — cols O-P, rows 4-15 ───────────────────────────
hdr_main(ws, 4, 15, 16, "  RESULTS SUMMARY", bg=T_SECHDR, sz=10)
hdr_col(ws, 5, 15, "Metric",  bg=T_COLHDR)
hdr_col(ws, 5, 16, "Value",   bg=T_COLHDR)

# Dynamic INDEX references — adjust offsets for new row layout
# gen inputs now start at row 6 (not 5): n_ext=$B$6, n_scr=$B$7, etc.
# Update REF to match new row layout
REF.update({
    "n_ext": "$B$6",  "n_scr": "$B$7",  "n_str": "$B$8",
    "He":    "$B$9",  "Hs":    "$B$10", "Ht":    "$B$11",
    "F":     "$B$12", "S":     "$B$13",
    "O":     "$B$14", "R":     "$B$15",
    "P_Pr":  "$H$6",  "Q_Pr":  "$I$6",  "XF_Pr": "$M$6",
    "P_Nd":  "$H$7",  "Q_Nd":  "$I$7",  "XF_Nd": "$M$7",
    "P_Tb":  "$H$8",  "Q_Tb":  "$I$8",  "XF_Tb": "$M$8",
    "P_Dy":  "$H$9",  "Q_Dy":  "$I$9",  "XF_Dy": "$M$9",
})

ne, ns = REF["n_ext"], REF["n_scr"]
lo_row   = f"19+{ne}"       # SOLVER row of stage n_ext  (R0-1+n = 19+n)
preg_idx = f"20+{ne}+{ns}"  # SOLVER row of first strip stage
raff_row = R0               # stage 1 always at R0 = row 20

xcl = {e: gcl(XCOL[e]) for e in ELEMS}
def idx(col_ltr, row_expr): return f"INDEX(${col_ltr}:${col_ltr},{row_expr})"
def raff(e):  return f"${xcl[e]}${raff_row}"
def preg(e):  return idx(xcl[e], preg_idx)
sum_feed = "+".join(REF[f"XF_{e}"] for e in ELEMS)
sum_raff = "+".join(raff(e) for e in ELEMS)
sum_preg = "+".join(preg(e) for e in ELEMS)

results = [
    ("Loaded Organic @ Extraction Exit (g/L)", f"=INDEX($AM:$AM,{lo_row})","0.00"),
    ("Raffinate — Nd  (g/L)",   f"={raff('Nd')}", "0.00"),
    ("Raffinate — Dy  (g/L)",   f"={raff('Dy')}", "0.00"),
    ("Preg — Nd  (g/L)",        f"={preg('Nd')}", "0.00"),
    ("Preg — Dy  (g/L)",        f"={preg('Dy')}", "0.00"),
    ("% Feed  —  Nd", f"=IF(({sum_feed})=0,0,{REF['XF_Nd']}/({sum_feed})*100)","0.00"),
    ("% Feed  —  Dy", f"=IF(({sum_feed})=0,0,{REF['XF_Dy']}/({sum_feed})*100)","0.00"),
    ("% Purity in Raffinate  —  Nd",
     f"=IF(({sum_raff})=0,0,{raff('Nd')}/({sum_raff})*100)","0.00"),
    ("% Purity in Preg  —  Dy",
     f"=IF(({sum_preg})=0,0,{preg('Dy')}/({sum_preg})*100)","0.00"),
    ("Scrub Flowrate  =  S × R",     f"={REF['S']}*{REF['R']}","0.00"),
    ("Preg Flowrate   =  S × (1−R)", f"={REF['S']}*(1-{REF['R']})","0.00"),
]
for i,(lbl,frm,nf) in enumerate(results, start=6):
    label(ws, i, 15, lbl, bg=O_BG, fc=O_FG)
    c=ws.cell(row=i,column=16); style_output(c, frm=frm, nf=nf, sz=10)

# ── Panel E: O/A Ratio Tracker — cols R-T (18-20), rows 4-10 ─────────────
hdr_main(ws, 4, 18, 20, "  O/A RATIO TRACKER", bg=T_SECHDR, sz=10)
hdr_col(ws, 5, 18, "Section",  bg=T_COLHDR)
hdr_col(ws, 5, 19, "O/A",      bg=T_COLHDR)
hdr_col(ws, 5, 20, "Status",   bg=T_COLHDR)

_F = REF["F"]   # "$B$12"  feed to extraction
_S = REF["S"]   # "$B$13"  scrub (= preg bleed flowrate)
_O = REF["O"]   # "$B$14"  organic flowrate
_R = REF["R"]   # "$B$15"  reflux ratio  (SR = S*R  goes back as scrub)

# SR = scrub flowrate returned to circuit = S * R
# Extraction: aqueous in = SR + F
# Scrub:      aqueous in = SR
# Strip:      aqueous in = S  (strip acid)
oa_rows = [
    ("Extraction",
     f"=IF(({_S}*{_R}+{_F})=0,0,{_O}/({_S}*{_R}+{_F}))"),
    ("Scrub",
     f"=IF(({_S}*{_R})=0,0,{_O}/({_S}*{_R}))"),
    ("Strip",
     f"=IF({_S}=0,0,{_O}/{_S})"),
]
for i,(sec_lbl, oa_frm) in enumerate(oa_rows, start=6):
    label(ws, i, 18, sec_lbl, bg=I_BG, fc=I_FG)
    oa_cell = ws.cell(row=i, column=19)
    style_output(oa_cell, frm=oa_frm, nf="0.00", sz=10)
    oa_ref = f"S{i}"
    st_cell = ws.cell(row=i, column=20)
    st_cell.value = f'=IF({oa_ref}=0,"—",IF({oa_ref}<0.5,"LOW — flood risk",IF({oa_ref}>5,"HIGH — underloading","OK")))'
    st_cell.font      = font(col=O_FG, sz=9, bold=True)
    st_cell.fill      = fill(O_BG)
    st_cell.alignment = aln(h="center")
    st_cell.border    = brd()

# ═════════════════════════════════════════════════════════════════════════════
# SOLVER HEADER ROWS 17-19
# ═════════════════════════════════════════════════════════════════════════════
ws.row_dimensions[16].height = 6

# Row 17: SOLVER banner
ws.merge_cells(f"A17:{gcl(40)}17")
c = ws["A17"]
c.value = ("SOLVER  ·  TDMA (Thomas Algorithm)  ·  "
           "Columns:  D_i  |  a  b  c  d  (matrix)  |  c′  d′  (forward sweep →)  |  "
           "y_org  (back substitution ←)  |  x_aq  (aqueous output)")
c.font  = font(bold=True, col=WHT, sz=10)
c.fill  = fill(T_SECHDR)
c.alignment = aln(h="left")
ws.row_dimensions[17].height = 20

# Row 18: element group banners
ws.merge_cells("A18:B18")
ws["A18"].fill = fill(T_SECHDR)
for e in ELEMS:
    s = ESTART[e]
    ws.merge_cells(f"{gcl(s)}18:{gcl(s+7)}18")
    c = ws[f"{gcl(s)}18"]
    c.value = f"{e}  (8 columns)"
    c.font  = font(bold=True, col=WHT, sz=10)
    c.fill  = fill(E_CLR[e])
    c.alignment = aln()
ws.merge_cells(f"{gcl(35)}18:{gcl(38)}18")
c=ws[f"{gcl(35)}18"]; c.value="Aqueous  x_i"; c.font=font(bold=True,col=WHT,sz=10); c.fill=fill(T_SECHDR); c.alignment=aln()
ws.merge_cells(f"{gcl(39)}18:{gcl(40)}18")
c=ws[f"{gcl(39)}18"]; c.value="Totals"; c.font=font(bold=True,col=WHT,sz=10); c.fill=fill(T_SECHDR); c.alignment=aln()
ws.row_dimensions[18].height = 18

# Row 19: column headers with colour showing cell type
sub = {0:("D_i",C_BG),1:("a",C_BG),2:("b",C_BG),3:("c",C_BG),4:("d (RHS)",C_BG),
       5:("c′ →",S_BG),6:("d′ →",S_BG),7:("y_org ←",O_BG)}
hdr_col(ws, 19, 1, "Stage",   bg=T_COLHDR, sz=9)
hdr_col(ws, 19, 2, "Section", bg=T_COLHDR, sz=9)
for e in ELEMS:
    s = ESTART[e]
    for off,(lbl,cbg) in sub.items():
        c = ws.cell(row=19, column=s+off)
        c.value = lbl
        c.font  = font(bold=True, col=T_TITLE if cbg==C_BG else (O_FG if cbg==O_BG else S_FG), sz=8)
        c.fill  = fill(cbg)
        c.alignment = aln(wrap=True)
        c.border = brd("AAAAAA")
for off,lbl in enumerate(["x_Pr","x_Nd","x_Tb","x_Dy"]):
    c=ws.cell(row=19,column=35+off); c.value=lbl
    c.font=font(bold=True,col=R_FG,sz=8); c.fill=fill(R_BG); c.alignment=aln(); c.border=brd("AAAAAA")
for col,lbl in [(39,"Σ Organic"),(40,"Σ Aqueous")]:
    c=ws.cell(row=19,column=col); c.value=lbl
    c.font=font(bold=True,col=O_FG,sz=8); c.fill=fill(O_BG); c.alignment=aln(); c.border=brd("AAAAAA")
ws.row_dimensions[19].height = 26

# ═════════════════════════════════════════════════════════════════════════════
# SOLVER DATA ROWS 20-44
# ═════════════════════════════════════════════════════════════════════════════
def sec_bg(n):
    if n<=10: return SEC_EXT
    elif n<=20: return SEC_SCR
    else: return SEC_STR

for n in range(1, N+1):
    row = R0 + n - 1
    sbg = sec_bg(n)

    # Stage# and Section label
    c=ws.cell(row=row,column=1); c.value=n
    c.font=font(bold=True,col=T_TITLE,sz=10); c.fill=fill(sbg); c.alignment=aln(); c.border=brd()
    c=ws.cell(row=row,column=2)
    c.value=(f'=IF(A{row}<={REF["n_ext"]},"Extraction",'
             f'IF(A{row}<={REF["n_ext"]}+{REF["n_scr"]},"Scrub","Strip"))')
    c.font=font(col=T_TITLE,sz=9); c.fill=fill(sbg); c.alignment=aln(); c.border=brd()

    for e in ELEMS:
        s   = ESTART[e]
        Dc  = gcl(s);   ac  = gcl(s+1); bc  = gcl(s+2)
        cc  = gcl(s+3); dc  = gcl(s+4); cpc = gcl(s+5)
        dpc = gcl(s+6); yc  = gcl(s+7)

        frms = [f_D(row,e), f_a(row), f_b(row,Dc), f_c(row,Dc), f_d(row,e),
                f_cp(row,bc,cc,ac,cpc), f_dp(row,bc,dc,ac,cpc,dpc), f_y(row,dpc,cpc,yc)]
        styles = [C_BG,C_BG,C_BG,C_BG,C_BG, S_BG,S_BG, O_BG]
        fgs    = [C_FG,C_FG,C_FG,C_FG,C_FG, S_FG,S_FG, O_FG]
        bolds  = [False]*5 + [False,False,True]

        for off,(frm,bg,fg,bd) in enumerate(zip(frms,styles,fgs,bolds)):
            c = ws.cell(row=row, column=s+off)
            c.value = frm
            c.font  = font(bold=bd, col=fg, sz=9)
            c.fill  = fill(bg)
            c.alignment = aln()
            c.border = brd()
            c.number_format = "0.0000"

    # x_i = y/D  (REF style)
    for e in ELEMS:
        s=ESTART[e]; Dc=gcl(s); yc=gcl(s+7)
        c=ws.cell(row=row,column=XCOL[e])
        c.value=f"=IF({Dc}{row}<>0,{yc}{row}/{Dc}{row},0)"
        c.font=font(col=R_FG,sz=9); c.fill=fill(R_BG)
        c.alignment=aln(); c.border=brd(); c.number_format="0.0000"

    # Total organic (OUTPUT)
    c=ws.cell(row=row,column=39)
    c.value="="+"+".join(f"{gcl(ESTART[e]+7)}{row}" for e in ELEMS)
    c.font=font(bold=True,col=O_FG,sz=9); c.fill=fill(O_BG)
    c.alignment=aln(); c.border=brd(); c.number_format="0.0000"

    # Total aqueous (OUTPUT)
    c=ws.cell(row=row,column=40)
    c.value="="+"+".join(f"{gcl(XCOL[e])}{row}" for e in ELEMS)
    c.font=font(bold=True,col=O_FG,sz=9); c.fill=fill(O_BG)
    c.alignment=aln(); c.border=brd(); c.number_format="0.0000"

# Conditional formatting — section tints on cols A:B
cf = f"A{R0}:B{R1}"
ws.conditional_formatting.add(cf, FormulaRule(
    formula=[f"$A{R0}>{REF['n_ext']}+{REF['n_scr']}"],
    fill=PatternFill("solid",fgColor=SEC_STR), stopIfTrue=False))
ws.conditional_formatting.add(cf, FormulaRule(
    formula=[f"$A{R0}>{REF['n_ext']}"],
    fill=PatternFill("solid",fgColor=SEC_SCR), stopIfTrue=False))
ws.conditional_formatting.add(cf, FormulaRule(
    formula=[f"$A{R0}<={REF['n_ext']}"],
    fill=PatternFill("solid",fgColor=SEC_EXT), stopIfTrue=False))

# ═════════════════════════════════════════════════════════════════════════════
# STAGE PROFILE TABLE (for charts)
# ═════════════════════════════════════════════════════════════════════════════
ws.row_dimensions[R1+2].height = 6

hdr_main(ws, PROF_HDR, 1, 9,
         "  STAGE PROFILE  ·  Aqueous concentrations (g/L)  ·  Source for charts",
         bg=T_SECHDR, sz=10)
ws.row_dimensions[PROF_HDR].height = 20

for col,lbl in enumerate(["Stage","Section","x_Pr","x_Nd","x_Tb","x_Dy",
                           "x_NdPr\n(Nd+Pr)","%_NdPr","%_Dy"],start=1):
    hdr_col(ws, PROF_HDR+1, col, lbl, bg=T_COLHDR, sz=9, wrap=True)
ws.row_dimensions[PROF_HDR+1].height = 28

def sec_label(n):
    if n<=10: return "Extraction", SEC_EXT
    elif n<=20: return "Scrub",    SEC_SCR
    else: return "Strip",          SEC_STR

for n in range(1,N+1):
    pr = PROF_R0 + n - 1
    sr = R0      + n - 1
    sec, sbg = sec_label(n)

    c=ws.cell(row=pr,column=1); c.value=n
    c.font=font(col=T_TITLE,sz=9); c.fill=fill(sbg); c.alignment=aln(); c.border=brd()
    c=ws.cell(row=pr,column=2); c.value=sec
    c.font=font(col=T_TITLE,sz=9); c.fill=fill(sbg); c.alignment=aln(); c.border=brd()

    for j,e in enumerate(ELEMS,start=3):
        c=ws.cell(row=pr,column=j)
        style_ref(c, frm=f"=${gcl(XCOL[e])}${sr}", nf="0.0000")

    # x_NdPr
    c=ws.cell(row=pr,column=7)
    style_ref(c, frm=f"=C{pr}+D{pr}", nf="0.0000")

    tot=f"(C{pr}+D{pr}+E{pr}+F{pr})"
    c=ws.cell(row=pr,column=8)
    style_output(c, frm=f"=IF({tot}=0,0,(C{pr}+D{pr})/{tot}*100)", nf="0.00", sz=9)
    c=ws.cell(row=pr,column=9)
    style_output(c, frm=f"=IF({tot}=0,0,F{pr}/{tot}*100)", nf="0.00", sz=9)

# ═════════════════════════════════════════════════════════════════════════════
# CHARTS
# ═════════════════════════════════════════════════════════════════════════════
cats = Reference(ws, min_col=1, max_col=1, min_row=PROF_R0, max_row=PROF_R1)

def make_chart(title, y_title, col1, col2, y_min=None, y_max=None, pct_fmt=False):
    ch = LineChart()
    ch.title  = title
    ch.style  = 2        # clean minimal style
    ch.width  = 22
    ch.height = 14
    ch.x_axis.title    = "Stage"
    ch.y_axis.title    = y_title
    ch.x_axis.numFmt   = "0"
    ch.y_axis.numFmt   = '0"%"' if pct_fmt else "0.00"
    ch.x_axis.tickLblPos = "low"
    ch.legend.position = "tr"
    if y_min is not None: ch.y_axis.scaling.min = y_min
    if y_max is not None: ch.y_axis.scaling.max = y_max

    for col, color, label_text in [(col1,"1B3A6B",None),(col2,"C07B30",None)]:
        s = Series(Reference(ws, min_col=col, max_col=col,
                             min_row=PROF_HDR+1, max_row=PROF_R1),
                   title_from_data=True)
        s.graphicalProperties.line.solidFill = color
        s.graphicalProperties.line.width     = 22000
        s.marker.symbol = "circle"
        s.marker.size   = 5
        s.marker.graphicalProperties.solidFill   = color
        s.marker.graphicalProperties.line.solidFill = color
        ch.append(s)

    ch.set_categories(cats)
    return ch

ch1 = make_chart("Aqueous Distributions — Absolute",
                 "Aqueous Concentration  X  (g/L)", col1=7, col2=6)
ch2 = make_chart("Aqueous Distributions — % of Total",
                 "Aqueous Distribution  (%)",
                 col1=8, col2=9, y_min=0, y_max=100, pct_fmt=True)

ws.add_chart(ch1, f"A{PROF_R1+3}")
ws.add_chart(ch2, f"M{PROF_R1+3}")

# ═════════════════════════════════════════════════════════════════════════════
# SAPONIFICATION  &  ORGANIC LOADING
# Placed well below charts (charts float; data rows below are safe)
# ═════════════════════════════════════════════════════════════════════════════
from openpyxl.worksheet.datavalidation import DataValidation

SAP_START = PROF_R1 + 42       # first row of this section (≈ row 114)

# ── Molecular weights & constants ─────────────────────────────────────────────
MW = {"Pr": 140.91, "Nd": 144.24, "Tb": 158.93, "Dy": 162.50}
MW_D2EHPA   = 322.4   # g/mol
MW_CYN801   = 290.4   # g/mol  (bis(2,4,4-trimethylpentyl)phosphinic acid)
DEN_D2EHPA  = 0.975   # g/mL
DEN_CYN801  = 0.870   # g/mL
DEN_ORFOM   = 0.785   # g/mL
DEN_KERO    = 0.800   # g/mL
MW_NaOH     = 40.0

# y_org column indices: Pr=J(10), Nd=R(18), Tb=Z(26), Dy=AH(34)
YCOL_IDX = {"Pr": 10, "Nd": 18, "Tb": 26, "Dy": 34}
YCOL_LET = {"Pr": "J", "Nd": "R", "Tb": "Z", "Dy": "AH"}
# Total organic column = AM (39)
AM_COL = 39

ne = REF["n_ext"]   # "$B$5"
O  = REF["O"]       # "$B$13"

# ── Helpers ───────────────────────────────────────────────────────────────────
def sap_lbl(row, col, txt, bold=False, span=None):
    if span:
        ws.merge_cells(f"{gcl(col)}{row}:{gcl(col+span-1)}{row}")
    c = ws.cell(row=row, column=col)
    c.value = txt
    c.font  = font(bold=bold, col="2E4057", sz=9)
    c.fill  = fill("F5F7FA")
    c.alignment = aln(h="left")
    c.border = brd()

def sap_unit(row, col, txt):
    c = ws.cell(row=row, column=col)
    c.value = txt
    c.font  = font(col="757575", sz=8)
    c.fill  = fill("F5F7FA")
    c.alignment = aln(h="left")
    c.border = brd()

def sap_blank(row, col, span=1):
    for cc in range(col, col+span):
        c = ws.cell(row=row, column=cc)
        c.fill = fill("F5F7FA")
        c.border = brd()

def panel_hdr(row, c1, c2, txt, bg=T_SECHDR):
    ws.merge_cells(f"{gcl(c1)}{row}:{gcl(c2)}{row}")
    c = ws.cell(row=row, column=c1)
    c.value = txt
    c.font  = font(bold=True, col=WHT, sz=9)
    c.fill  = fill(bg)
    c.alignment = aln(h="left")
    c.border = brd("AAAAAA")
    ws.row_dimensions[row].height = 17

def sub_hdr(row, c1, c2, txt, bg="4A6FA5"):
    ws.merge_cells(f"{gcl(c1)}{row}:{gcl(c2)}{row}")
    c = ws.cell(row=row, column=c1)
    c.value = txt
    c.font  = font(bold=True, col=WHT, sz=8)
    c.fill  = fill(bg)
    c.alignment = aln(h="left")
    c.border = brd("AAAAAA")
    ws.row_dimensions[row].height = 15

# ── Section main header ────────────────────────────────────────────────────────
ws.row_dimensions[SAP_START - 1].height = 12   # spacer
ws.merge_cells(f"A{SAP_START}:{gcl(14)}{SAP_START}")
c = ws.cell(row=SAP_START, column=1)
c.value = "  SAPONIFICATION  ·  ORGANIC LOADING  ·  REAGENT REQUIREMENTS"
c.font  = font(bold=True, col=WHT, sz=12)
c.fill  = fill(T_TITLE)
c.alignment = aln(h="left")
ws.row_dimensions[SAP_START].height = 26

# ── Panel sub-headers (row SAP_START+1) ───────────────────────────────────────
panel_hdr(SAP_START+1,  1,  4, "  Organic System — Inputs")
panel_hdr(SAP_START+1,  6,  9, "  Saponification Calculation")
panel_hdr(SAP_START+1, 11, 14, "  Maximum Organic Loading")

# ── Column headers (row SAP_START+2) ──────────────────────────────────────────
ws.row_dimensions[SAP_START+2].height = 15
for col, txt, bg in [
        (1,"Parameter",T_COLHDR),(2,"Value",T_COLHDR),(3,"Units",T_COLHDR),(4,"",T_COLHDR),
        (6,"Parameter",T_COLHDR),(7,"Value",T_COLHDR),(8,"Units",T_COLHDR),(9,"",T_COLHDR),
        (11,"Parameter",T_COLHDR),(12,"Value",T_COLHDR),(13,"Units",T_COLHDR),(14,"",T_COLHDR),
]:
    hdr_col(ws, SAP_START+2, col, txt, bg=bg, sz=8)
sap_blank(SAP_START+2, 5); sap_blank(SAP_START+2, 10)

# ── Data rows start ────────────────────────────────────────────────────────────
r = SAP_START + 3    # first data row

# Set uniform row height for data rows
for i in range(16):
    ws.row_dimensions[r + i].height = 16

# ──────────────────────────────────────────────────────────────────────────────
# PANEL 1 — Organic System Inputs  (cols 1-4)
# ──────────────────────────────────────────────────────────────────────────────
p1_rows = [
    # (label,                 col_B_kind,  default_val,   fmt,       units        )
    ("Extractant",            "input_str", "Cyanex 801",  None,      "—"          ),
    ("Extractant  MW",        "calc",      None,          "0.0",     "g/mol"      ),
    ("Diluent",               "input_str", "Orfom",       None,      "—"          ),
    ("Diluent density",       "calc",      None,          "0.000",   "g/mL"       ),
    ("Organic density",       "calc",      None,          "0.000",   "g/mL"       ),
    ("% Extractant",          "input_num", 10.0,          "0.0",     "vol %"      ),
    ("[Extractant]",          "calc",      None,          "0.0",     "g/L org"    ),
    ("[Extractant]",          "calc",      None,          "0.000",   "mol/L org"  ),
    ("NaOH concentration",    "input_num", 20.0,          "0.0",     "wt %"       ),
    ("NaOH soln density",     "calc",      None,          "0.000",   "g/mL"       ),
    ("[NaOH] solution",       "calc",      None,          "0.000",   "mol/L"      ),
    ("NaOH flowrate",         "input_num", 0.0,           "0.000",   "same unit as O"),
]

for i, (lbl, kind, defval, fmt, units) in enumerate(p1_rows):
    ro = r + i
    sap_lbl(ro, 1, lbl)
    sap_blank(ro, 4)
    sap_unit(ro, 3, units)
    c = ws.cell(row=ro, column=2)
    c.border = brd()
    c.alignment = aln()
    if kind == "input_str":
        style_input(c, val=defval)
    elif kind == "input_num":
        style_input(c, val=defval, nf=fmt)
    else:  # calc — style only, formula set below
        c.font  = font(col=C_FG, sz=9)
        c.fill  = fill(C_BG)
        if fmt: c.number_format = fmt

# Calc formulas for Panel 1
# r+0: extractant (input) ; r+1: MW extractant
ws.cell(row=r+1,  column=2).value  = f'=IF(B{r}="D2EHPA",{MW_D2EHPA},IF(B{r}="Cyanex 801",{MW_CYN801},{MW_D2EHPA}))'
ws.cell(row=r+1,  column=2).number_format = "0.0"
# r+2: diluent (input) ; r+3: diluent density
ws.cell(row=r+3,  column=2).value  = f'=IF(B{r+2}="Orfom",{DEN_ORFOM},IF(B{r+2}="Kerosene",{DEN_KERO},{DEN_ORFOM}))'
ws.cell(row=r+3,  column=2).number_format = "0.000"
# r+4: organic density (vol-weighted)
ws.cell(row=r+4,  column=2).value  = (f'=B{r+5}/100'
                                       f'*IF(B{r}="D2EHPA",{DEN_D2EHPA},{DEN_CYN801})'
                                       f'+(1-B{r+5}/100)*B{r+3}')
ws.cell(row=r+4,  column=2).number_format = "0.000"
# r+6: [extractant] g/L (vol% × organic density)
ws.cell(row=r+6,  column=2).value  = f'=B{r+5}/100*B{r+4}*1000'
ws.cell(row=r+6,  column=2).number_format = "0.0"
# r+7: [extractant] mol/L
ws.cell(row=r+7,  column=2).value  = f'=IF(B{r+1}=0,0,B{r+6}/B{r+1})'
ws.cell(row=r+7,  column=2).number_format = "0.000"
# r+9: NaOH solution density approx  (ρ ≈ 1 + 0.011×wt%)
ws.cell(row=r+9,  column=2).value  = f'=1+0.011*B{r+8}'
ws.cell(row=r+9,  column=2).number_format = "0.000"
# r+10: [NaOH] mol/L
ws.cell(row=r+10, column=2).value  = f'=B{r+8}/100*B{r+9}*1000/{MW_NaOH}'
ws.cell(row=r+10, column=2).number_format = "0.000"

# Dropdowns
dv_ext = DataValidation(type="list", formula1='"D2EHPA,Cyanex 801"',
                        allow_blank=False, showDropDown=False)
dv_dil = DataValidation(type="list", formula1='"Orfom,Kerosene"',
                        allow_blank=False, showDropDown=False)
ws.add_data_validation(dv_ext)
ws.add_data_validation(dv_dil)
dv_ext.add(ws.cell(row=r,   column=2))
dv_dil.add(ws.cell(row=r+2, column=2))

# ──────────────────────────────────────────────────────────────────────────────
# PANEL 2 — Saponification  (cols 6-9)
# ──────────────────────────────────────────────────────────────────────────────
# Sub-block A: loaded organic at extraction exit
sub_hdr(r, 6, 8, "  Loaded Organic at Extraction Exit")
ws.row_dimensions[r].height = 15

for i, (elem, mw) in enumerate(MW.items(), start=1):
    ro = r + i
    col_let = YCOL_LET[elem]
    sap_lbl(ro, 6, f"y_{elem}  (org. conc., ext exit)")
    c = ws.cell(row=ro, column=7)
    style_ref(c, frm=f"=INDEX(${col_let}:${col_let},19+{ne})", nf="0.0000")
    sap_unit(ro, 8, "g/L org")
    sap_blank(ro, 9)

# r+5: total
sap_lbl(r+5, 6, "Total loaded organic", bold=True)
c = ws.cell(row=r+5, column=7)
style_output(c, frm=f"=G{r+1}+G{r+2}+G{r+3}+G{r+4}", nf="0.000", sz=9)
sap_unit(r+5, 8, "g/L org")
sap_blank(r+5, 9)

# Sub-block B: Theoretical (Alind Eq.10 adapted)
sub_hdr(r+6, 6, 8, "  Theoretical NaOH  (Alind Eq. 10 — 3 mol NaOH per mol REE³⁺)")
ws.row_dimensions[r+6].height = 15

sap_lbl(r+7, 6,  "NaOH required")
sap_lbl(r+8, 6,  "NaOH solution flowrate")
sap_lbl(r+9, 6,  "Pure NaOH mass rate")

# mol NaOH/time = 3 × O × Σ(y_i / MW_i)
c = ws.cell(row=r+7, column=7)
style_calc(c, nf="0.000", sz=9,
           frm=(f"=3*{O}*(G{r+1}/{MW['Pr']}"
                f"+G{r+2}/{MW['Nd']}"
                f"+G{r+3}/{MW['Tb']}"
                f"+G{r+4}/{MW['Dy']})"))
sap_unit(r+7, 8, "mol / [O·time]")
sap_blank(r+7, 9)

# Volume of NaOH solution / time
c = ws.cell(row=r+8, column=7)
style_calc(c, nf="0.000", sz=9, frm=f"=IF(B{r+10}=0,0,G{r+7}/B{r+10})")
sap_unit(r+8, 8, "L soln / [O·time]")
sap_blank(r+8, 9)

# Pure NaOH mass / time
c = ws.cell(row=r+9, column=7)
style_calc(c, nf="0.000", sz=9, frm=f"=G{r+7}*{MW_NaOH}/1000")
sap_unit(r+9, 8, "kg NaOH / [O·time]")
sap_blank(r+9, 9)

# Sub-block C: Experimental saponification degree
sub_hdr(r+10, 6, 8, "  Experimental Saponification Degree")
ws.row_dimensions[r+10].height = 15

sap_lbl(r+11, 6, "mol NaOH supplied  [per time unit]")
sap_lbl(r+12, 6, "mol D2EHPA/extr. in organic  [per time unit]")
sap_lbl(r+13, 6, "SAPONIFICATION DEGREE", bold=True)

c = ws.cell(row=r+11, column=7)
style_calc(c, nf="0.000", sz=9, frm=f"=B{r+11}*B{r+10}")
sap_unit(r+11, 8, "mol / time")
sap_blank(r+11, 9)

c = ws.cell(row=r+12, column=7)
style_calc(c, nf="0.000", sz=9, frm=f"=B{r+7}*{O}")
sap_unit(r+12, 8, "mol / time")
sap_blank(r+12, 9)

c = ws.cell(row=r+13, column=7)
style_output(c, nf="0.0", sz=10,
             frm=f"=IF(G{r+12}=0,0,G{r+11}/G{r+12}*100)")
sap_unit(r+13, 8, "% saponified")
sap_blank(r+13, 9)

# ──────────────────────────────────────────────────────────────────────────────
# PANEL 3 — Maximum Organic Loading  (cols 11-14)
# ──────────────────────────────────────────────────────────────────────────────
sub_hdr(r, 11, 13, "  Extractant Capacity")
ws.row_dimensions[r].height = 15   # already set above; OK to re-confirm

sap_lbl(r+1, 11, "Available capacity  [extractant / 3]")
c = ws.cell(row=r+1, column=12)
style_calc(c, nf="0.000", sz=9, frm=f"=B{r+7}/3")
sap_unit(r+1, 13, "mol REE / L org")
sap_blank(r+1, 14)

sub_hdr(r+2, 11, 13, "  Max Loading per Element  (if pure feed)")
ws.row_dimensions[r+2].height = 15

for i, (elem, mw) in enumerate(MW.items(), start=3):
    ro = r + i
    sap_lbl(ro, 11, f"Max loading — {elem}  (pure {elem} feed)")
    c = ws.cell(row=ro, column=12)
    style_calc(c, nf="0.0", sz=9, frm=f"=L{r+1}*{mw}")
    sap_unit(ro, 13, "g/L org")
    sap_blank(ro, 14)

sub_hdr(r+7, 11, 13, "  Max Loading — Current Feed Composition")
ws.row_dimensions[r+7].height = 15

# Feed-weighted average REE MW
xf = [REF[f"XF_{e}"] for e in ELEMS]
mws_list = list(MW.values())
num_wt = "+".join(f"{xf[i]}*{mws_list[i]}" for i in range(4))
den_wt  = "+".join(xf)
sap_lbl(r+8, 11, "Avg REE MW  (feed-weighted)")
c = ws.cell(row=r+8, column=12)
style_calc(c, nf="0.00", sz=9,
           frm=f"=IF(({den_wt})=0,144.24,({num_wt})/({den_wt}))")
sap_unit(r+8, 13, "g/mol")
sap_blank(r+8, 14)

sap_lbl(r+9, 11, "Max loading — mixed REE", bold=True)
c = ws.cell(row=r+9, column=12)
style_output(c, nf="0.0", sz=10, frm=f"=L{r+1}*L{r+8}")
sap_unit(r+9, 13, "g/L org")
sap_blank(r+9, 14)

sap_lbl(r+10, 11, "Current loaded organic  (from SOLVER)")
c = ws.cell(row=r+10, column=12)
style_ref(c, nf="0.0", frm=f"=INDEX($AM:$AM,19+{ne})")
sap_unit(r+10, 13, "g/L org")
sap_blank(r+10, 14)

ws.row_dimensions[r+10].height = 16

sap_lbl(r+11, 11, "Loading utilization", bold=True)
c = ws.cell(row=r+11, column=12)
style_output(c, nf="0.0", sz=10,
             frm=f"=IF(L{r+9}=0,0,L{r+10}/L{r+9}*100)")
sap_unit(r+11, 13, "%  of max capacity used")
sap_blank(r+11, 14)

ws.row_dimensions[r+11].height = 16

# Spacer col between panels
for row in range(SAP_START, r+16):
    ws.cell(row=row, column=5).fill  = fill("FFFFFF")
    ws.cell(row=row, column=10).fill = fill("FFFFFF")

# ═════════════════════════════════════════════════════════════════════════════
# RESIDENCE TIME  (settler dimensions)
# ═════════════════════════════════════════════════════════════════════════════
RT_START = r + 18          # first row of Residence Time section

ws.row_dimensions[RT_START - 1].height = 12   # spacer
ws.merge_cells(f"A{RT_START}:{gcl(14)}{RT_START}")
c = ws.cell(row=RT_START, column=1)
c.value = "  RESIDENCE TIME  ·  Settler Dimensions"
c.font  = font(bold=True, col=WHT, sz=11)
c.fill  = fill(T_SECHDR)
c.alignment = aln(h="left")
ws.row_dimensions[RT_START].height = 22

# ── Mixer & Settler inputs ─────────────────────────────────────────────────
sub_hdr(RT_START+1, 1, 14, "Mixer / Settler Inputs")
rt_inp = [
    ("Mixer active volume",    0.10,  "0.000",  "m³  per stage"),
    ("Settler length",         2.0,   "0.00",   "m"),
    ("Settler width",          1.0,   "0.00",   "m"),
    ("Settler depth (active)", 0.30,  "0.000",  "m"),
    ("No. settlers per stage", 1,     "0",      "—"),
]
rt_inp_rows = {}     # label → absolute row
for ii, (lbl, defval, fmt, units) in enumerate(rt_inp):
    ro = RT_START + 2 + ii
    rt_inp_rows[lbl] = ro
    sap_lbl(ro, 1, lbl)
    c = ws.cell(row=ro, column=2)
    style_input(c, val=defval, nf=fmt)
    sap_unit(ro, 3, units)
    sap_blank(ro, 4)

# Convenience absolute refs
rv  = f"$B${RT_START+2}"   # mixer volume m³
rl  = f"$B${RT_START+3}"   # settler length
rw  = f"$B${RT_START+4}"   # settler width
rd  = f"$B${RT_START+5}"   # settler depth
rns = f"$B${RT_START+6}"   # num settlers per stage

# Aqueous flows per section (L/min  =  same units as O)
# Ext:   Q_aq = S*R + F   (scrub reflux + feed)
# Scr:   Q_aq = S*R
# Str:   Q_aq = S
_ne = REF["n_ext"]; _ns = REF["n_scr"]; _nt = REF["n_str"]
_F2 = REF["F"]; _S2 = REF["S"]; _O2 = REF["O"]; _R2 = REF["R"]

sub_hdr(RT_START+8, 1, 14, "Residence Time per Section")
rt_tbl_hdr_row = RT_START + 9
for col, lbl in [(1,"Section"),(2,"n stages"),(3,"Q_aq"),(4,"Q_tot"),
                  (5,"τ_mixer (min)"),(6,"V_settler (m³)"),(7,"τ_settler (min)"),(8,"τ_stage (min)"),(9,"τ_section (min)")]:
    hdr_col(ws, rt_tbl_hdr_row, col, lbl, bg=T_COLHDR, sz=8)

rt_sec_rows = {}   # section → absolute row  (for SS section to reference)
for jj, (sec_name, n_ref, qa_frm) in enumerate([
    ("Extraction", _ne, f"{_S2}*{_R2}+{_F2}"),
    ("Scrub",      _ns, f"{_S2}*{_R2}"),
    ("Strip",      _nt, f"{_S2}"),
]):
    ro = RT_START + 10 + jj
    rt_sec_rows[sec_name] = ro
    sap_lbl(ro, 1, sec_name)
    # n stages
    c = ws.cell(row=ro, column=2); style_ref(c, frm=f"={n_ref}", nf="0")
    # Q_aq
    c = ws.cell(row=ro, column=3); style_calc(c, frm=f"={qa_frm}", nf="0.00")
    # Q_tot = O + Q_aq
    c = ws.cell(row=ro, column=4); style_calc(c, frm=f"={_O2}+C{ro}", nf="0.00")
    # τ_mixer = V_mixer(L) / Q_tot  (V in m³ → ×1000 for L, Q already in L/min assumed)
    c = ws.cell(row=ro, column=5); style_calc(c, frm=f"=IF(D{ro}=0,0,{rv}*1000/D{ro})", nf="0.00")
    # V_settler = L × W × D × n_settlers
    c = ws.cell(row=ro, column=6); style_calc(c, frm=f"={rl}*{rw}*{rd}*{rns}", nf="0.000")
    # τ_settler (min) = V_settler(m³)*1000 / Q_tot
    c = ws.cell(row=ro, column=7); style_calc(c, frm=f"=IF(D{ro}=0,0,F{ro}*1000/D{ro})", nf="0.00")
    # τ_stage = τ_mixer + τ_settler
    c = ws.cell(row=ro, column=8); style_calc(c, frm=f"=E{ro}+G{ro}", nf="0.00")
    # τ_section = n_stages × τ_stage
    c = ws.cell(row=ro, column=9); style_output(c, frm=f"=B{ro}*H{ro}", nf="0.00")

# ═════════════════════════════════════════════════════════════════════════════
# STEADY-STATE TIME ESTIMATOR
# Ref: Ritcey & Ashbrook, "Solvent Extraction: Principles and Applications to
#      Process Metallurgy" (1979), Vol.II §6; Perry's Chemical Engineers'
#      Handbook 8th Ed., §15-42 (mixer-settler turnover); control theory
#      heuristic 3τ ≈ 95% of steady state.
# ═════════════════════════════════════════════════════════════════════════════
SS_START = RT_START + 15

ws.row_dimensions[SS_START - 1].height = 12
ws.merge_cells(f"A{SS_START}:{gcl(14)}{SS_START}")
c = ws.cell(row=SS_START, column=1)
c.value = ("  STEADY-STATE TIME ESTIMATOR"
           "  ·  Ritcey & Ashbrook (1979) §6  ·  Perry's 8th Ed. §15-42  ·  3τ heuristic")
c.font  = font(bold=True, col=WHT, sz=10)
c.fill  = fill(T_SECHDR)
c.alignment = aln(h="left")
ws.row_dimensions[SS_START].height = 22

sub_hdr(SS_START+1, 1, 9, "Inputs")
sap_lbl(SS_START+2, 1, "Safety / turnover factor  (default = 3, i.e., 3τ ≈ 95% SS)")
c = ws.cell(row=SS_START+2, column=2); style_input(c, val=3, nf="0")
sap_unit(SS_START+2, 3, "—")

sub_hdr(SS_START+4, 1, 9, "Section Residence Times  (from Residence Time section above)")
for col, lbl in [(1,"Section"),(2,"τ_section (min)"),(3,"τ_section (h)")]:
    hdr_col(ws, SS_START+5, col, lbl, bg=T_COLHDR, sz=8)

ss_tau_refs = []
for jj, sec_name in enumerate(["Extraction","Scrub","Strip"]):
    ro = SS_START + 6 + jj
    src_ro = rt_sec_rows[sec_name]
    sap_lbl(ro, 1, sec_name)
    c = ws.cell(row=ro, column=2); style_ref(c, frm=f"=I{src_ro}", nf="0.00")
    c = ws.cell(row=ro, column=3); style_calc(c, frm=f"=B{ro}/60", nf="0.000")
    ss_tau_refs.append(f"B{ro}")

tau_cells = ",".join(ss_tau_refs)
max_row = SS_START + 10

sap_lbl(max_row,   1, "τ_max  (bottleneck section)")
c = ws.cell(row=max_row, column=2); style_output(c, frm=f"=MAX({tau_cells})", nf="0.00")
sap_unit(max_row, 3, "min")

sap_lbl(max_row+1, 1, "Bottleneck section")
c = ws.cell(row=max_row+1, column=2)
c.value = (f'=IF(B{max_row}=B{SS_START+6},"Extraction",'
           f'IF(B{max_row}=B{SS_START+7},"Scrub","Strip"))')
c.font  = font(col=O_FG, sz=9, bold=True)
c.fill  = fill(O_BG)
c.alignment = aln(h="left")
c.border = brd()

sap_lbl(max_row+2, 1, "t_SS  =  factor × τ_max", bold=True)
c = ws.cell(row=max_row+2, column=2)
style_output(c, frm=f"=$B${SS_START+2}*B{max_row}", nf="0.00")
sap_unit(max_row+2, 3, "min")
c2 = ws.cell(row=max_row+2, column=4)
style_calc(c2, frm=f"=B{max_row+2}/60", nf="0.00")
sap_unit(max_row+2, 5, "h")

# ═════════════════════════════════════════════════════════════════════════════
# FEED VARIABILITY SCENARIO  (linear scaling — Dy only)
# Since TDMA is linear in feed conc (d = -XF×F), scaling XF_Dy by k
# scales all Dy outputs by exactly k.  Other elements are unchanged.
# ═════════════════════════════════════════════════════════════════════════════
FV_START = SS_START + 18

ws.row_dimensions[FV_START - 1].height = 12
ws.merge_cells(f"A{FV_START}:{gcl(14)}{FV_START}")
c = ws.cell(row=FV_START, column=1)
c.value = "  FEED VARIABILITY SCENARIO  ·  Dy ± scaling  ·  Linear TDMA property"
c.font  = font(bold=True, col=WHT, sz=11)
c.fill  = fill(T_SECHDR)
c.alignment = aln(h="left")
ws.row_dimensions[FV_START].height = 22

sub_hdr(FV_START+1, 1, 14, "Scenario Multipliers  (applied to Dy feed concentration only)")
sap_lbl(FV_START+2, 1, "k1  (high Dy scenario, e.g. Dy + 20%)")
c = ws.cell(row=FV_START+2, column=2); style_input(c, val=1.20, nf="0.00")
sap_unit(FV_START+2, 3, "×  XF_Dy")
sap_lbl(FV_START+3, 1, "k2  (low  Dy scenario, e.g. Dy − 20%)")
c = ws.cell(row=FV_START+3, column=2); style_input(c, val=0.80, nf="0.00")
sap_unit(FV_START+3, 3, "×  XF_Dy")

k1_ref = f"$B${FV_START+2}"
k2_ref = f"$B${FV_START+3}"

# Column headers for 3-column comparison table
fv_hdr_row = FV_START + 5
for col, lbl in [(1,"Metric"),(2,"Normal"),(5,"Dy + k1"),(8,"Dy + k2")]:
    hdr_col(ws, fv_hdr_row, col, lbl, bg=T_COLHDR, sz=9)

# Base SOLVER expressions (from existing REF)
_preg_nd = preg("Nd")   # INDEX formula for Nd preg
_preg_dy = preg("Dy")
_raff_nd = raff("Nd")
_raff_dy = raff("Dy")
_lo      = f"INDEX($AM:$AM,{lo_row})"

fv_metrics = [
    ("Loaded Organic (g/L)",     _lo,      _lo,         _lo),
    ("Raffinate  —  Nd (g/L)",   _raff_nd, _raff_nd,    _raff_nd),
    ("Raffinate  —  Dy (g/L)",   _raff_dy, f"{_raff_dy}*{k1_ref}", f"{_raff_dy}*{k2_ref}"),
    ("Preg  —  Nd (g/L)",        _preg_nd, _preg_nd,    _preg_nd),
    ("Preg  —  Dy (g/L)",        _preg_dy, f"{_preg_dy}*{k1_ref}", f"{_preg_dy}*{k2_ref}"),
    ("Preg purity  —  Dy  (%)",
     f"=IF(({sum_preg})=0,0,{_preg_dy}/({sum_preg})*100)",
     None, None),   # purity handled specially below
    ("Raff purity  —  Nd  (%)",
     f"=IF(({sum_raff})=0,0,{_raff_nd}/({sum_raff})*100)",
     None, None),
]

for jj, (lbl, base, hi, lo_val) in enumerate(fv_metrics):
    ro = FV_START + 6 + jj
    sap_lbl(ro, 1, lbl)
    # Normal column (cols 2-4 merged for readability)
    if base.startswith("="):
        frm_n = base
    else:
        frm_n = f"={base}"
    c = ws.cell(row=ro, column=2); style_output(c, frm=frm_n, nf="0.00")

    # Purity rows: scale numerator and denominator separately
    if lbl.startswith("Preg purity"):
        # Normal
        # Hi: Dy × k1, rest unchanged → sum_preg denominator also changes for Dy
        _sp_hi = f"({sum_preg}-{_preg_dy}+{_preg_dy}*{k1_ref})"
        _sp_lo = f"({sum_preg}-{_preg_dy}+{_preg_dy}*{k2_ref})"
        frm_hi = f"=IF({_sp_hi}=0,0,{_preg_dy}*{k1_ref}/{_sp_hi}*100)"
        frm_lo_v = f"=IF({_sp_lo}=0,0,{_preg_dy}*{k2_ref}/{_sp_lo}*100)"
    elif lbl.startswith("Raff purity"):
        _sr_hi = f"({sum_raff}-{_raff_dy}+{_raff_dy}*{k1_ref})"
        _sr_lo = f"({sum_raff}-{_raff_dy}+{_raff_dy}*{k2_ref})"
        frm_hi = f"=IF({_sr_hi}=0,0,{_raff_nd}/{_sr_hi}*100)"
        frm_lo_v = f"=IF({_sr_lo}=0,0,{_raff_nd}/{_sr_lo}*100)"
    else:
        frm_hi  = f"={hi}"  if hi  else frm_n
        frm_lo_v = f"={lo_val}" if lo_val else frm_n

    c = ws.cell(row=ro, column=5); style_output(c, frm=frm_hi,   nf="0.00")
    c = ws.cell(row=ro, column=8); style_output(c, frm=frm_lo_v, nf="0.00")

# Spacer cols between scenario columns
for ro in range(FV_START+5, FV_START+14):
    for cc in [3,4,6,7,9]:
        ws.cell(row=ro, column=cc).fill = fill("F5F7FA")

# ── Tab colour ────────────────────────────────────────────────────────────────
ws.sheet_properties.tabColor = "1B3A6B"

# ── Save ──────────────────────────────────────────────────────────────────────
out = "/home/user/SX-simulation/SX_Steady_State_Model.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Layout:  legend row 2 | inputs rows 4-15 | solver rows 17-44 | profile rows {PROF_HDR}-{PROF_R1} | charts rows {PROF_R1+3}+ | saponification rows {SAP_START}+")
